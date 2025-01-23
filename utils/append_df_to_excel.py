def append_df_to_excel(data, sheet, save_path):
    import pandas as pd
    from openpyxl import load_workbook
    import os
    from openpyxl import Workbook

    # Check if the output file exists. Create one if it does not
    if os.path.exists(save_path):
        # Open the file as a workbook in read-only mode
        workbook = load_workbook(save_path)   # open an Excel file

        # Check if sheet exists; create one if it does not
        if sheet not in workbook.sheetnames:
            workbook.create_sheet(sheet)
            workbook.save(save_path)

        # Read existing data from the sheet into a DataFrame
        existing_data = pd.read_excel(save_path, sheet_name=sheet)

        # Append the new data to the existing DataFrame
        updated_data = pd.concat([existing_data, data], ignore_index=True)

        # Overwrite the existing sheet with the updated DataFrame
        with pd.ExcelWriter(
            save_path,
            engine='openpyxl',
            mode='a',
            if_sheet_exists='replace'  # Ensure existing sheet is replaced
        ) as writer:
            updated_data.to_excel(writer, sheet_name=sheet, index=False, float_format="%6f")
            # writer.save()
    else:
        # Create a new file and write to it
        with pd.ExcelWriter(
            save_path,
            engine='openpyxl',
            mode='w'
        ) as writer:
            data.to_excel(writer, sheet_name=sheet, index=False, float_format="%6f")
    